import requests
from bs4 import BeautifulSoup
import openpyxl

wb = openpyxl.Workbook();
sheet = wb.active

dictionary = {}

for i in range(1, 4):
    res = requests.get('https://search.naver.com/search.naver?where=news&sm=tab_jum&query=2018+여의도+불꽃축제&start=' + str(i * 10 + 1), headers={'User-Agent': 'Mozilla/5.0'})
    raw = res.text
    html = BeautifulSoup(raw, 'html.parser')

    articles = html.select('ul.type01 > li')

    for article in articles:
        title = article.select_one('a._sp_each_title').text
        journal = article.select_one('span._sp_each_source').text
        print(title, '/', journal)


        if journal in dictionary.keys():
            dictionary[journal].append(title)
        else:
            dictionary[journal] = [title]

    print('---------', i, '페이지 수집중', '----------')

for d in dictionary.items():
    print(d)

row = 1
for d in dictionary.items():
    sheet.cell(row = row, column=1).value = d[0]

    for item in d[1]:
        sheet.cell(row = row, column = 2).value = item
        row +=1
    row +=1

wb.save('naver_news_firework.xlsx')
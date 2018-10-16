
import openpyxl

wb = openpyxl.Workbook()


sheet1 = wb.active
sheet1.title = 'Openpyxl 연습'

sheet1['A1'] = 'Hello Openpyxl'
sheet1.append(['일요일','월요일','화요일','수요일','목요일','금요일','토요일'])

sheet2 = wb.create_sheet('새 시트')
sheet2.title = '계단식 배치'

for i in range(10):

    sheet2.cell(row=i+1, column= i+1).value = str(i+1)

wb.save('challenge.xlsx')
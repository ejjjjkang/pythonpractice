# ---- 1. 엑셀 파일 만들고 저장하기 ----
#
# import openpyxl
#
# wb = openpyxl.Workbook()
# wb.save('text.xlsx')


# ---- 2. 엑셀 시트 & 셀에 접근하고 수정하기 ----

# import openpyxl
#
# wb = openpyxl.Workbook()
# sheet = wb.active
#
# sheet['D5'] = 'hello world'
# sheet.cell(row=2,column=2).value= '3, 3'
# sheet.append([1,2,3,4,5])
#
#
# wb.save('test2.xlsx')


# ---- 3. 시트 변환 ----

# import openpyxl
#
# wb = openpyxl.Workbook()
# sheet1 = wb.active
# sheet1.title = '수집 데이터'
# sheet1['A1'] = '첫번째 시트'
#
# sheet2 = wb.create_sheet('새 시트')
# sheet2['A1'] = '두번째 시트'
#
# sheet1['A2'] = '다시 첫번째 시트'
#
# wb.save('test3.xlsx')


# ---- 4. 기존 파일 불러오기 ----

import openpyxl

wb = openpyxl.load_workbook('test2.xlsx')

# sheet1 = wb.active
sheet1 = wb['새 시트']


sheet1.title = "이름 변경"
sheet1.append(range(10))

wb.save('test2.xlsx')